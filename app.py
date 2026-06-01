from shiny import App, ui, render, reactive, req
from shinyswatch import theme
import pandas as pd
import plotly.express as px
import tempfile
import os
import json
from google.oauth2 import service_account
from google.cloud import bigquery
from dotenv import load_dotenv
import uuid
from pathlib import Path
from data import queries, processing


# #Authenticate
# key_path = Path(__file__).parent / "recruitment_bq.json"
# credentials = service_account.Credentials.from_service_account_file(str(key_path))
# client = bigquery.Client(credentials=credentials)
# client = queries.get_client()

# Load environment variables from .env file
load_dotenv()
# Authenticate
# Use individual environment variables
service_account_info = {
    "type": "service_account",
    "project_id": "rugbaleeg",
    "private_key_id": os.getenv("GCP_PRIVATE_KEY_ID"),
    "private_key": os.getenv("GCP_PRIVATE_KEY").replace('\\n', '\n'),  # Fix newlines
    "client_email": os.getenv("GCP_CLIENT_EMAIL"),
    "client_id": os.getenv("GCP_CLIENT_ID"),
    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
    "token_uri": "https://oauth2.googleapis.com/token",
    "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
    "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/shiny-recruitment%40rugbaleeg.iam.gserviceaccount.com",
    "universe_domain": "googleapis.com"
}
credentials = service_account.Credentials.from_service_account_info(service_account_info)
client = bigquery.Client(credentials=credentials)
contracts_df = queries.fetch_bq_contract_data(client)

TEMPLATE_FILE = Path(__file__).parent / "player_table_templates.json"


def load_templates():
    if TEMPLATE_FILE.exists():
        with open(TEMPLATE_FILE, "r") as f:
            return json.load(f)
    return {}


def save_templates(templates):
    with open(TEMPLATE_FILE, "w") as f:
        json.dump(templates, f, indent=2)


CONTRACT_END_COLORS = {
    2026: "#FF4444",  # red
    2027: "#E6970A",  # amber
}

broncos_maroon = '#540c2b'
broncos_gold = '#f0a91f'



comps_dict = {
    111:'nrl',
    113:'nswcup',
    114:'qcup',
    116:'origin',
    121:'superleague',
    143:'haroldmatthews',
    144:'sgball',
    153:'nrlq',
    154:'cyrilconnell',
    155:'malmeninga',
    159:'jerseyflegg',
    161:'nrlw'
}

seasons_list = [2023,2024,2025,2026]

summary_list = ['Game Average', 'Game Totals', 'Per 80 Mins', 'Individual Games']

positions_list = ['Fullback', 'Winger', 'Centre', 'Five-Eighth', 'Halfback',
                  'Hooker','Prop', '2nd Row', 'Lock', 'Interchange']

with open("stats.json", "r") as f:
    stats_dict =json.load(f)

stats_flattened_dict = {k: v for d in stats_dict.values() for k, v in d.items()}

default_stats = ['Rating','allRuns','allRunMetres',
                 'tries','tryAssists',
                 'linebreaks','linebreakAssists',
                 'tackleBreaks',
                 'tackles','missedTackles','ineffectiveTackles','effectiveTacklePercentage',
                 'errors','penalties']



def eval_rule(value, op, threshold):
    if pd.isna(value): return False
    if op == ">":  return value > threshold
    if op == "<":  return value < threshold
    if op == ">=": return value >= threshold
    if op == "<=": return value <= threshold
    if op == "=":  return abs(value - threshold) < 0.001
    return False

def hex_to_rgba(hex_color, alpha):
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


leaderboard_defaults = {
    'WG': ['tries','allRuns','tackles'],
    'CT': ['tries','allRuns','tackles'],
    'FB': ['tries','allRuns','tackles'],
    'FE': ['tries','allRuns','tackles'],
    'HB': ['tries','allRuns','tackles'],
    'HK': ['tries','allRuns','tackles'],
    '2R': ['tries','allRuns','tackles'],
    'PR': ['tries','allRuns','tackles'],
    'LK': ['tries','allRuns','tackles'],
    'INT': ['tries','allRuns','tackles']
}

ranking_position_groups = [
    'Halves', 'Hookers', 'Starting Middles', 
    'Backrowers', 'Outside Backs', 'Bench'
]

ranking_methods = {
    'percentile_rank': 'Percentile Rank',
    'zscore': 'Z-Score',
    'minmax': 'Min-Max'
}

ranking_defaults = {
    'Halves':           ['tries_per80', 'tryAssists_per80', 'linebreaks_per80', 'kickMetres_per80', 'errors_per80'],
    'Hookers':          ['tries_per80', 'tackleBreaks_per80', 'runs_per80', 'tackles_per80', 'errors_per80'],
    'Starting Middles': ['metres_per80', 'runs_per80', 'postContactMetres_per80', 'tackleAttempts_per80', 'errors_per80'],
    'Backrowers':       ['metres_per80', 'tackleBreaks_per80', 'offloads_per80', 'tackleAttempts_per80', 'errors_per80'],
    'Outside Backs':    ['tries_per80', 'metres_per80', 'tackleBreaks_per80', 'kickDefusalPct', 'errors_per80'],
    'Bench':            ['metres_per80', 'runs_per80', 'tackleAttempts_per80', 'errors_per80', 'sixAgains_per80'],
}

all_ranking_metrics = [
    'tries_per80', 'tryAssists_per80', 'linebreaks_per80', 'linebreakAssists_per80',
    'tackleBreaks_per80', 'runs_per80', 'metres_per80', 'postContactMetres_per80',
    'offloads_per80', 'receipts_per80', 'supports_per80', 'halfbreaks_per80',
    'tackleAttempts_per80', 'errors_per80', 'penalties_per80', 'sixAgains_per80',
    'tackleEfficiency', 'ptbWinPct', 'postContactMetresPct', 'metresPerRun',
    'postContactMetresPerRun', 'runsPerTackleBreak', 'kickReturnMetresperkickReturn',
    'passesPerRun', 'receiptsPerError', 'kickDefusalPct', 'fastPtbPct', 'goalKickingPct'
]

def create_position_tabs():
    tabs = []
    for position_abbrev, default_stats in leaderboard_defaults.items():
        tab = ui.nav_panel(
            position_abbrev,  # Display full name
            ui.input_selectize(
                f"stats_{position_abbrev}",  # Use code as ID
                None,
                choices={category:stats_dict[category] for category in ['Derived','Attack','Defence','Discipline','Kicking']},
                selected=default_stats,
                multiple=True,
                width="100%"
            ),
            ui.output_ui(f"cards_{position_abbrev}")
        )
        tabs.append(tab)
    return tabs


def create_ranking_tabs():
    tabs = []
    for group in ranking_position_groups:
        tab = ui.nav_panel(
            group,
            ui.input_selectize(
                f"ranking_metrics_{group.replace(' ', '_')}",
                "Metrics:",
                choices=all_ranking_metrics,
                selected=ranking_defaults[group],
                multiple=True,
                width="100%"
            ),
            ui.output_data_frame(f"ranking_table_{group.replace(' ', '_')}")
        )
        tabs.append(tab)
    return tabs

#UI

#Add pages

#Home will be about games completed/data collected
#Player Table
#Leaderboards
#Analyzer file downloader + video player

home_page = ui.nav_panel(
    "Home",
    ui.h2("Welcome to the Recruitment Dashboard"),
    ui.p("Select a page from the navigation above. See the up to date data below."),
    ui.output_ui("fixture_cards")
)

table_page = ui.nav_panel(
    "Player Table",
    ui.h2("Player Statistics Table"),
    ui.p("Use the filters on the left to customise the player statistics table. You can sort by a column by clicking the header."),
    ui.layout_sidebar(
        ui.sidebar(
            ui.h4("Templates"),

            ui.input_select(
                "player_table_template",
                "Template:",
                choices=[],
                selected=None,
                width="100%"
            ),

            ui.div(
                {"style": "display:flex; gap:6px; margin-bottom:8px;"},
                ui.input_action_button(
                    "apply_player_table_template",
                    "Apply",
                    class_="btn btn-sm btn-outline-primary"
                ),
                ui.input_action_button(
                    "delete_player_table_template",
                    "Delete",
                    class_="btn btn-sm btn-outline-danger"
                ),
            ),

            ui.input_text(
                "new_template_name",
                "Save current view as:",
                placeholder="e.g. Middle forwards"
            ),

            ui.input_action_button(
                "save_player_table_template",
                "Save template",
                class_="btn btn-sm btn-outline-success"
            ),

            ui.hr(style="margin-top: 8px; margin-bottom: 8px;"),
            ui.h4("Filters"),
            ui.input_selectize("summary",
                            "Summary Type:",
                            choices=summary_list,
                            selected='Game Average'),
            ui.input_slider("min_games", 
                            "Minimum Games:", 
                            1, 10, 1, step=1),
            ui.input_checkbox_group("game_types",
                                    None, 
                                    choices=['Regular', 'Finals'],
                                    selected=['Regular'],
                                    inline=True),
            ui.hr(style="margin-top: 0px; margin-bottom: 0px;"),
            ui.input_selectize("team",
                            "Teams:",
                            choices=[],
                            selected=[],
                            multiple=True),
            ui.input_selectize("player",
                            "Players:",
                            choices=[],
                            selected=[],
                            multiple=True),
            ui.input_selectize("position",
                            "Positions:",
                            choices=positions_list,
                            selected=[],
                            multiple=True),
            ui.input_checkbox("position_separate",
                            "Separate Positions", 
                            value=False),
            ui.input_checkbox("season_separate",
                            "Separate Seasons", 
                            value=False),
            ui.input_checkbox("comp_separate",
                            "Separate Comps", 
                            value=False),
            ui.hr(style="margin-top: 0px; margin-bottom: 0px;"),
            ui.input_selectize("stats",
                            "Stats:",
                            choices={category:stats_dict[category] for category in ['Derived','Attack','Defence','Discipline','Kicking']},
                            selected=default_stats,
                            multiple=True),
            ui.hr(style="margin-top: 0px; margin-bottom: 0px;"),
            ui.h6("Highlight Rules"),
            ui.output_ui("highlight_rules_ui"),
            ui.input_action_button("add_rule", "Add rule", class_="btn btn-sm btn-outline-secondary mt-1"),
            width=450
        ),
        # Custom CSS
        ui.tags.style("""
            #player_table table {
                width: max-content !important;
                table-layout: auto !important;
            }

            /* Never wrap names */
            #player_table table th:first-child,
            #player_table table td:first-child {
                width: 180px !important;
                min-width: 180px !important;
                max-width: 180px !important;

                white-space: nowrap !important;
                overflow: hidden !important;
                text-overflow: ellipsis !important;

                padding-left: 4px !important;
                padding-right: 12px !important;
            }

            /* Compact numeric columns */
            #player_table table th:not(:first-child),
            #player_table table td:not(:first-child) {
                white-space: nowrap !important;
                width: 1% !important;
                padding-left: 8px !important;
                padding-right: 8px !important;
                text-align: center !important;
            }

            /* Header cells */
            #player_table table thead th:not(:first-child) {
                height: 100px !important;
                padding-top: 20px !important;
                padding-bottom: 10px !important;
                vertical-align: bottom !important;
                text-align: center !important;
            }

            /* Rotated header text */
            #player_table table thead th:not(:first-child) div,
            #player_table table thead th:not(:first-child) span {
                writing-mode: vertical-rl !important;
                transform: rotate(180deg) !important;

                display: flex !important;

                justify-content: flex-start !important;
                align-items: center !important;

                margin: 0 auto !important;

                white-space: nowrap !important;
            }
        """),
        ui.download_button(
            "download_player_table",
            "Download Excel",
            class_="btn btn-sm btn-outline-success"
        ),
        ui.output_data_frame("player_table")
    )
)

leaderboard_page = ui.nav_panel(
    "Leaderboards",
    ui.h2("Leaderboards"),
    ui.p("See top players in various statistical categories."),
    ui.layout_columns(
        ui.input_selectize("leaderboard_summary",
            "Summary Type:",
            choices=['Game Average', 'Game Totals','Game Best'],
            selected='Game Average'),
        ui.input_slider("leaderboard_min_games", 
            "Minimum Games:", 
            value=5,
            min=1,
            max=10,
            step=1),
        ui.input_slider("leaderboard_top_n",
            "Top N Players:",
            value=5,
            min=1,
            max=10,
            step=1),
        col_widths=[4, 4, 4],  # Each takes up 1/3 of the width
        style="margin-bottom: 1rem;max-height: 50px;" 
    ),
    ui.navset_tab(
        *create_position_tabs(),
    )
)

rankings_page = ui.nav_panel(
    "Season Rankings",
    ui.h2("Season Rankings"),
    ui.p("Season-to-date rankings normalised per 80 minutes. Minimum 200 minutes played."),
    ui.layout_columns(
        ui.input_select(
            "ranking_method",
            "Ranking Method:",
            choices=ranking_methods,
            selected='percentile_rank',
            width="150px"
        ),
        col_widths=[4],
        style="margin-bottom: 1rem;max-height: 50px;"
    ),
    ui.navset_tab(
        *create_ranking_tabs()
    )
)

app_ui = ui.page_navbar(
    home_page,
    table_page,
    leaderboard_page,
    rankings_page,
    ui.nav_spacer(),
    ui.nav_control(
        ui.div(
            ui.input_selectize("competition",
                None,
                choices={str(i):j for i,j in comps_dict.items()},
                selected=['111'],
                multiple=True,
                width='200px'
            ),
        )        
    ),
    ui.nav_control(
        ui.div(
            ui.input_selectize("season",
                None,
                choices=[str(i) for i in seasons_list],
                selected=['2026'],
                multiple=True,
                width='200px'
            ),
        )
    ),
    title="Recruitment Dashboard",
    fillable=True,
    theme=theme.flatly(),
    id="navbar",
    header=ui.tags.link(href="css.css", rel="stylesheet")
)

#Server
def server(input, output, session):

    player_table_templates = reactive.value(load_templates())

    @reactive.effect
    def update_template_choices():
        templates = player_table_templates()

        ui.update_select(
            "player_table_template",
            choices=list(templates.keys()),
            selected=input.player_table_template()
        )

    @reactive.calc
    def fixture_data():
        reactive.invalidate_later(7200) 
        df = queries.fetch_bq_latest_fixtures(client)
        return df.sort_values(by=['competitionName','roundId','gameNumber'])

    @render.ui
    def fixture_cards():
        df = fixture_data()
        if len(df) == 0 or df is None:
            return ui.p("No fixtures available.")
        
        grouped = df.groupby(['competitionId','competitionName','roundName'])

        cards = []
        for (comp_id, comp_name, round_name), group_df in grouped:
            table_df = group_df[['gameNumber','game','gameStateName']].copy()
            table_df['gameStateName'] = table_df['gameStateName'].apply(
                lambda x: '✅' if x == 'Final' else '❌'
            )
            table_html = table_df.to_html(
                index=False,
                header=False, 
                classes="table table-hover table-sm fixture-columns",
                border=0
            )
            cards.append(
                ui.card(
                    ui.card_header(f"{comp_name} - {round_name}"),
                    ui.HTML(table_html)
                )
            )
        
        return ui.div(*cards)

    @reactive.calc
    def bigquery_data():
        reactive.invalidate_later(86400)
        comps = [int(c) for c in input.competition()]
        seasons = [int(s) for s in input.season()]
        req(comps, seasons)

        stats = [k for c in stats_dict.keys() if c != "Derived" for k in stats_dict[c]]
        df = queries.fetch_bq_player_data(client, comps, seasons, stats)
        df = processing.calculate_rating(df)
        return df

    @reactive.calc
    def rankings_data():
        reactive.invalidate_later(86400)
        comps = [int(c) for c in input.competition()]
        seasons = [int(s) for s in input.season()]
        req(comps, seasons)

        return queries.fetch_bq_rankings_data(client, comps, seasons)

    @reactive.effect
    def update_team_choices():
        df = bigquery_data()
        teams = (
            df[["teamId", "teamNickName"]]
            .drop_duplicates()
            .sort_values("teamNickName")
            .set_index("teamId")["teamNickName"]
            .astype(str)
            .to_dict()
        )
        teams = {str(k): v for k, v in teams.items()}
        ui.update_select("team", choices=teams)

    @reactive.effect
    def update_player_choices():
        df = bigquery_data()
        selected_teams = input.team()

        if selected_teams:
            df = df[df["teamId"].astype(str).isin(selected_teams)]

        players = (
            df[["playerId", "playerName"]]
            .drop_duplicates()
            .sort_values("playerName") 
            .set_index("playerId")["playerName"]
            .astype(str)
            .to_dict()
        )
        players = {str(k): v for k, v in players.items()}
        ui.update_select("player", choices=players)

    @reactive.calc
    def filtered_data():
        df = bigquery_data()

        game_types = input.game_types()

        teams = input.team()
        teams = [int(t) for t in teams] if teams else None

        players = input.player()
        players = [int(p) for p in players] if players else None

        positions = input.position()

        stats = list(input.stats())
        stats = list(stats_dict['Always'].keys()) + stats

        filtered_df = processing.filter_bq_player_data(
            df, game_types, teams, players, positions, stats
        )

        return filtered_df
    
    @reactive.calc
    def summarised_data():
        df = filtered_data()

        summary_type = input.summary()
        min_games = input.min_games()
        separate_positions = input.position_separate()
        separate_seasons = input.season_separate()
        separate_comps = input.comp_separate()

        stats = list(input.stats())
        stats = ['mins'] + stats

        summarised_df = processing.summarise_filtered_data(
            df,
            summary_type,
            min_games,
            separate_positions,
            separate_seasons,
            separate_comps,
            stats,
            stats_flattened_dict
        )

        summarised_df = processing.add_contract_info(summarised_df, contracts_df)

        return summarised_df

    @reactive.calc
    def display_table_data():
        df = summarised_data().reset_index(drop=True)
        display_df = df.drop(columns=["PID", "all_contract_end"], errors="ignore")
        return display_df

    # -------------------------
    # Highlight rules
    # -------------------------

    import uuid

    highlight_rules = reactive.value([])


    @reactive.effect
    @reactive.event(input.add_rule)
    def add_highlight_rule():
        display_df = display_table_data()
        numeric_cols = display_df.select_dtypes(include="number").columns.tolist()

        if not numeric_cols:
            return

        rules = list(highlight_rules())

        rules.append({
            "id": str(uuid.uuid4())[:8],
            "col": numeric_cols[0],
            "op": ">",
            "val": 0.0,
            "color": "#4ade80"
        })

        highlight_rules.set(rules)


    @render.ui
    def highlight_rules_ui():
        rules = highlight_rules()

        display_df = display_table_data()
        numeric_cols = display_df.select_dtypes(include="number").columns.tolist()

        if not numeric_cols:
            return ui.p("No numeric columns available to highlight.")

        rows = []

        for rule in rules:
            rid = rule["id"]

            rows.append(
                ui.div(
                    {
                        "style": (
                            "display:flex; gap:4px; align-items:center; "
                            "margin-bottom:4px;"
                        )
                    },

                    ui.input_select(
                        f"rule_col_{rid}",
                        None,
                        choices=numeric_cols,
                        selected=rule["col"] if rule["col"] in numeric_cols else numeric_cols[0],
                        width="140px"
                    ),

                    ui.input_select(
                        f"rule_op_{rid}",
                        None,
                        choices={
                            ">": ">",
                            "<": "<",
                            ">=": "≥",
                            "<=": "≤",
                            "=": "="
                        },
                        selected=rule["op"],
                        width="55px"
                    ),

                    ui.input_numeric(
                        f"rule_val_{rid}",
                        None,
                        value=rule["val"],
                        width="70px"
                    ),

                    ui.input_select(
                        f"rule_color_{rid}",
                        None,
                        choices={
                            "#4ade80": "Green",
                            "#facc15": "Amber",
                            "#f87171": "Red"
                        },
                        selected=rule["color"],
                        width="80px"
                    ),

                    ui.input_action_button(
                        f"rule_delete_{rid}",
                        "✕",
                        class_="btn btn-sm btn-outline-danger",
                        style="padding:2px 7px;"
                    ),
                )
            )

        return ui.div(*rows)


    @reactive.effect
    def sync_and_delete_highlight_rules():

        rules = list(highlight_rules())
        new_rules = []
        changed = False

        for rule in rules:

            rid = rule["id"]

            # -------------------------
            # Delete
            # -------------------------

            try:
                delete_count = input[f"rule_delete_{rid}"]()
            except Exception:
                delete_count = 0

            if delete_count > 0:
                changed = True
                continue

            # -------------------------
            # Sync values
            # -------------------------

            try:
                col = input[f"rule_col_{rid}"]()
                op = input[f"rule_op_{rid}"]()
                val = input[f"rule_val_{rid}"]()
                color = input[f"rule_color_{rid}"]()

            except Exception:
                new_rules.append(rule)
                continue

            if col is None or op is None or val is None or color is None:
                new_rules.append(rule)
                continue

            updated_rule = {
                "id": rid,
                "col": col,
                "op": op,
                "val": float(val),
                "color": color
            }

            if updated_rule != rule:
                changed = True

            new_rules.append(updated_rule)

        if changed:
            highlight_rules.set(new_rules)


    @reactive.calc
    def current_highlight_rules():
        return highlight_rules()


    # -------------------------
    # Player table templates
    # -------------------------

    @reactive.effect
    @reactive.event(input.save_player_table_template)
    def save_current_template():
        template_name = input.new_template_name()

        if not template_name:
            return

        templates = dict(player_table_templates())

        templates[template_name] = {
            "summary": input.summary(),
            "min_games": input.min_games(),
            "game_types": list(input.game_types()),
            "team": list(input.team()),
            "player": list(input.player()),
            "position": list(input.position()),
            "position_separate": input.position_separate(),
            "season_separate": input.season_separate(),
            "comp_separate": input.comp_separate(),
            "stats": list(input.stats()),
            "highlight_rules": [
                {
                    "col": rule["col"],
                    "op": rule["op"],
                    "val": rule["val"],
                    "color": rule["color"],
                }
                for rule in current_highlight_rules()
            ],
        }

        save_templates(templates)
        player_table_templates.set(templates)

        ui.update_select(
            "player_table_template",
            choices=list(templates.keys()),
            selected=template_name
        )

        ui.update_text("new_template_name", value="")


    @reactive.effect
    @reactive.event(input.apply_player_table_template)
    def apply_player_table_template():
        template_name = input.player_table_template()

        if not template_name:
            return

        templates = player_table_templates()

        if template_name not in templates:
            return

        template = templates[template_name]


        ui.update_selectize(
            "summary",
            selected=template.get("summary", "Game Average")
        )

        ui.update_slider(
            "min_games",
            value=template.get("min_games", 1)
        )

        ui.update_checkbox_group(
            "game_types",
            selected=template.get("game_types", ["Regular"])
        )

        ui.update_selectize(
            "team",
            selected=template.get("team", [])
        )

        ui.update_selectize(
            "player",
            selected=template.get("player", [])
        )

        ui.update_selectize(
            "position",
            selected=template.get("position", [])
        )

        ui.update_checkbox(
            "position_separate",
            value=template.get("position_separate", False)
        )

        ui.update_checkbox(
            "season_separate",
            value=template.get("season_separate", False)
        )

        ui.update_checkbox(
            "comp_separate",
            value=template.get("comp_separate", False)
        )

        ui.update_selectize(
            "stats",
            selected=template.get("stats", default_stats)
        )

        rules = []

        for rule in template.get("highlight_rules", []):
            rules.append({
                "id": str(uuid.uuid4())[:8],
                "col": rule["col"],
                "op": rule["op"],
                "val": rule["val"],
                "color": rule["color"],
            })

        highlight_rules.set(rules)


    @reactive.effect
    @reactive.event(input.delete_player_table_template)
    def delete_player_table_template():
        template_name = input.player_table_template()

        if not template_name:
            return

        templates = dict(player_table_templates())

        if template_name in templates:
            del templates[template_name]

        save_templates(templates)
        player_table_templates.set(templates)

        ui.update_select(
            "player_table_template",
            choices=list(templates.keys()),
            selected=None
        )


    # -------------------------
    # Player table
    # -------------------------

    @output
    @render.data_frame
    def player_table():
        df = summarised_data().reset_index(drop=True)
        display_df = display_table_data()

        styles = []

        for year, color in CONTRACT_END_COLORS.items():
            if "all_contract_end" in df.columns and "Name" in display_df.columns:
                matching_rows = df.index[df["all_contract_end"] == year].tolist()

                if matching_rows:
                    styles.append(render.StyleInfo(
                        rows=matching_rows,
                        cols=["Name"],
                        style={
                            "color": color,
                            "font-weight": "bold"
                        },
                    ))

        if "all_contract_end" in df.columns and "Name" in display_df.columns:
            unsigned_rows = df.index[df["all_contract_end"].isna()].tolist()

            if unsigned_rows:
                styles.append(render.StyleInfo(
                    rows=unsigned_rows,
                    cols=["Name"],
                    style={
                        "color": "#999999",
                        "font-weight": "bold"
                    },
                ))

        for rule in current_highlight_rules():

            col = rule["col"]

            if col not in display_df.columns:
                continue

            numeric_col = pd.to_numeric(display_df[col], errors="coerce")

            threshold = rule["val"]

            # -------------------------
            # Scale totals by matches
            # -------------------------

            if input.summary() in ["Game Totals"]:

                if "MAT" in display_df.columns:

                    row_threshold = (
                        pd.to_numeric(display_df["MAT"], errors="coerce")
                        * threshold
                    )

                else:
                    row_threshold = threshold

                mask = [
                    eval_rule(value, rule["op"], thresh)
                    for value, thresh in zip(numeric_col, row_threshold)
                ]

            else:

                mask = numeric_col.apply(
                    lambda x: eval_rule(x, rule["op"], threshold)
                )

            rows = display_df.index[mask].tolist()

            if rows:

                styles.append(
                    render.StyleInfo(
                        rows=rows,
                        cols=[col],
                        style={
                            "background-color": hex_to_rgba(rule["color"], 0.45),
                            "font-weight": "bold",
                        }
                    )
                )

        return render.DataTable(
            display_df,
            styles=styles,
            width="100%",
            height="99%"
        )


    # -------------------------
    # Download player table
    # -------------------------

    @render.download(filename="player_table.xlsx")
    def download_player_table():
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils.dataframe import dataframe_to_rows

        df = summarised_data().reset_index(drop=True)
        display_df = display_table_data()

        wb = Workbook()
        ws = wb.active
        ws.title = "Player Table"

        for row in dataframe_to_rows(display_df, index=False, header=True):
            ws.append(row)

        # Header styling
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Contract styling for Name column
        if "Name" in display_df.columns and "all_contract_end" in df.columns:
            name_col_idx = display_df.columns.get_loc("Name") + 1

            for row_idx, contract_end in enumerate(df["all_contract_end"], start=2):

                cell = ws.cell(row=row_idx, column=name_col_idx)

                if pd.isna(contract_end):
                    cell.font = Font(color="999999", bold=True)

                elif int(contract_end) == 2026:
                    cell.font = Font(color="FF4444", bold=True)

                elif int(contract_end) == 2027:
                    cell.font = Font(color="E6970A", bold=True)

        # Highlight rule styling
        for rule in current_highlight_rules():
            col = rule["col"]

            if col not in display_df.columns:
                continue

            col_idx = display_df.columns.get_loc(col) + 1
            numeric_col = pd.to_numeric(display_df[col], errors="coerce")
            threshold = rule["val"]

            if input.summary() in ["Game Totals", "Season Totals"] and "MAT" in display_df.columns:
                row_threshold = pd.to_numeric(display_df["MAT"], errors="coerce") * threshold

                mask = [
                    eval_rule(value, rule["op"], thresh)
                    for value, thresh in zip(numeric_col, row_threshold)
                ]
            else:
                mask = numeric_col.apply(
                    lambda x: eval_rule(x, rule["op"], threshold)
                )

            excel_color = rule["color"].replace("#", "").upper()
            fill = PatternFill(
                start_color=excel_color,
                end_color=excel_color,
                fill_type="solid"
            )

            for row_idx, should_highlight in enumerate(mask, start=2):
                if should_highlight:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.font = Font(bold=True)

        # Auto-width columns
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, len(str(value)))

            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


        output = BytesIO()
        wb.save(output)
        output.seek(0)

        yield output.getvalue()

    # -------------------------
    # Leaderboards
    # -------------------------

    def create_leaderboard_cards(position_abbrev):
        stats = input[f"stats_{position_abbrev}"]()

        if not stats:
            return ui.p("No stats selected.")
        
        summary_type = input.leaderboard_summary()
        cards = []

        for stat in stats:
            output_id = f"grid_{position_abbrev}_{stat}"
            stat_display_name = stats_flattened_dict.get(stat, stat)

            card = ui.div(
                {"class": "col-12 col-sm-6 col-md-4 col-lg-3"},
                ui.card(
                    ui.card_header(f"{stat_display_name} - {summary_type}"),
                    ui.output_data_frame(output_id),
                    class_="leaderboard-grid-card"
                )
            )
            cards.append(card)

            def make_renderer(pos_abbrev, stat_name, out_id):
                @output(id=out_id)
                @render.data_frame
                def _():
                    df = bigquery_data()
                    df = df[df['playerPositionAbbrev'] == pos_abbrev]

                    summary_type = input.leaderboard_summary()
                    min_games = input.leaderboard_min_games()
                    top_n = input.leaderboard_top_n()

                    leaderboard = processing.leaderboard_df(
                        df,
                        stat_name,
                        summary_type,
                        min_games,
                        top_n
                    )

                    return render.DataGrid(
                        leaderboard,
                        width="100%",
                        height="auto",
                        filters=False,
                        summary=False
                    )

                return _

            make_renderer(position_abbrev, stat, output_id)
        
        return ui.div({"class": "row"}, *cards)

    @output
    @render.ui
    def cards_WG():
        return create_leaderboard_cards('WG')
    
    @output
    @render.ui
    def cards_CT():
        return create_leaderboard_cards('CT')
    
    @output
    @render.ui
    def cards_FB():
        return create_leaderboard_cards('FB')
    
    @output
    @render.ui
    def cards_FE():
        return create_leaderboard_cards('FE')
    
    @output
    @render.ui
    def cards_HB():
        return create_leaderboard_cards('HB')
    
    @output
    @render.ui
    def cards_HK():
        return create_leaderboard_cards('HK')
    
    @output
    @render.ui
    def cards_2R():
        return create_leaderboard_cards('2R')
    
    @output
    @render.ui
    def cards_PR():
        return create_leaderboard_cards('PR')
    
    @output
    @render.ui
    def cards_LK():
        return create_leaderboard_cards('LK')
    
    @output
    @render.ui
    def cards_INT():
        return create_leaderboard_cards('INT')

    # -------------------------
    # Rankings
    # -------------------------

    for _group in ranking_position_groups:
        def make_ranking_renderer(g):
            @output(id=f"ranking_table_{g.replace(' ', '_')}")
            @render.data_frame
            def _():
                df = rankings_data()

                if df is None or len(df) == 0:
                    return render.DataGrid(pd.DataFrame())

                group_key = g.replace(' ', '_')
                selected_metrics = list(input[f"ranking_metrics_{group_key}"]())
                ranking_method = input.ranking_method()

                final_df = processing.pivot_rankings_data(
                    df,
                    g,
                    selected_metrics,
                    ranking_method
                )

                return render.DataGrid(
                    final_df,
                    width="100%",
                    filters=False,
                    summary=False
                )

            return _

        make_ranking_renderer(_group)


app_dir = Path(__file__).parent
app = App(app_ui, server, static_assets=app_dir / "www")
