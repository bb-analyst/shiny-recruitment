"""Player table page: filters, saved templates, highlight rules, Excel export."""

import uuid
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from shiny import module, reactive, render, ui
from shiny.module import resolve_id

import config
from data import processing, queries

# Operators offered in the highlight rule builder.
RULE_OPERATORS = {">": ">", "<": "<", ">=": "≥", "<=": "≤", "=": "="}

# Inline so the icon needs no extra dependency; currentColor makes it follow
# the button's own text colour on hover and focus.
DOWNLOAD_ICON = ui.HTML(
    '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" '
    'viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.5" '
    'stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">'
    '<path d="M8 2v8.5"/><path d="M4.5 7.5 8 11l3.5-3.5"/>'
    '<path d="M2.5 13.5h11"/></svg>'
)


def toolbar(*items):
    """A compact control strip above the table, right-aligned.

    Stands in for ui.toolbar, which is not available in this Shiny version.
    Add further controls as extra arguments; they lay out left to right.
    """
    return ui.div(
        {"class": "d-flex align-items-center justify-content-end gap-2 mb-2"},
        *items,
    )


def _table_css(table_id: str) -> str:
    """Column sizing and rotated headers for the player table.

    Takes the resolved (namespaced) output id so the selectors keep matching
    once the page is mounted as a module.
    """
    return f"""
        #{table_id} table {{
            width: max-content !important;
            table-layout: auto !important;
        }}

        /* Never wrap names */
        #{table_id} table th:first-child,
        #{table_id} table td:first-child {{
            width: 180px !important;
            min-width: 180px !important;
            max-width: 180px !important;

            white-space: nowrap !important;
            overflow: hidden !important;
            text-overflow: ellipsis !important;

            padding-left: 4px !important;
            padding-right: 12px !important;
        }}

        /* Compact numeric columns */
        #{table_id} table th:not(:first-child),
        #{table_id} table td:not(:first-child) {{
            white-space: nowrap !important;
            width: 1% !important;
            padding-left: 8px !important;
            padding-right: 8px !important;
            text-align: center !important;
        }}

        /* Header cells */
        #{table_id} table thead th:not(:first-child) {{
            height: 100px !important;
            padding-top: 20px !important;
            padding-bottom: 10px !important;
            vertical-align: bottom !important;
            text-align: center !important;
        }}

        /* Rotated header text */
        #{table_id} table thead th:not(:first-child) div,
        #{table_id} table thead th:not(:first-child) span {{
            writing-mode: vertical-rl !important;
            transform: rotate(180deg) !important;

            display: flex !important;

            justify-content: flex-start !important;
            align-items: center !important;

            margin: 0 auto !important;

            white-space: nowrap !important;
        }}
    """


def highlight_mask(display_df: pd.DataFrame, rule: dict, scale_by_matches: bool):
    """Rows in ``display_df`` that satisfy ``rule``.

    For cumulative summary types the threshold is per match, so it is scaled by
    the MAT column before comparing. Falls back to a flat threshold when MAT is
    not present.
    """
    values = pd.to_numeric(display_df[rule["col"]], errors="coerce")
    threshold = rule["val"]

    if scale_by_matches and "MAT" in display_df.columns:
        row_thresholds = pd.to_numeric(display_df["MAT"], errors="coerce") * threshold
        return [
            processing.eval_rule(value, rule["op"], row_threshold)
            for value, row_threshold in zip(values, row_thresholds)
        ]

    return [processing.eval_rule(value, rule["op"], threshold) for value in values]


def build_workbook(df: pd.DataFrame, display_df: pd.DataFrame, rules, scale_by_matches):
    """Render the player table to an Excel workbook, styling included."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Player Table"

    for row in dataframe_to_rows(display_df, index=False, header=True):
        ws.append(row)

    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Contract styling for Name column
    if "Name" in display_df.columns and "all_contract_end" in df.columns:
        name_col_idx = display_df.columns.get_loc("Name") + 1

        for row_idx, contract_end in enumerate(df["all_contract_end"], start=2):
            cell = ws.cell(row=row_idx, column=name_col_idx)

            if pd.isna(contract_end):
                cell.font = Font(color=config.UNSIGNED_COLOR.lstrip("#"), bold=True)
            elif int(contract_end) in config.CONTRACT_END_COLORS:
                color = config.CONTRACT_END_COLORS[int(contract_end)]
                cell.font = Font(color=color.lstrip("#"), bold=True)

    # Highlight rule styling
    for rule in rules:
        if rule["col"] not in display_df.columns:
            continue

        col_idx = display_df.columns.get_loc(rule["col"]) + 1
        mask = highlight_mask(display_df, rule, scale_by_matches)

        excel_color = rule["color"].replace("#", "").upper()
        fill = PatternFill(
            start_color=excel_color, end_color=excel_color, fill_type="solid"
        )

        for row_idx, should_highlight in enumerate(mask, start=2):
            if should_highlight:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.font = Font(bold=True)

    # Auto-width columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@module.ui
def player_table_ui(contract_end_choices):
    return ui.nav_panel(
        "Player Table",
        ui.layout_sidebar(
            ui.sidebar(
                ui.h4("Templates"),
                ui.input_selectize("template", "Template:", choices=[], selected=None, width="100%"),
                ui.div(
                    {"style": "display:flex; gap:6px; margin-bottom:8px;"},
                    ui.input_action_button(
                        "apply_template", "Apply", class_="btn btn-sm btn-outline-primary"
                    ),
                    ui.input_action_button(
                        "delete_template", "Delete", class_="btn btn-sm btn-outline-danger"
                    ),
                ),
                ui.input_text(
                    "new_template_name",
                    "Save current view as:",
                    placeholder="e.g. Middle forwards",
                ),
                ui.input_action_button(
                    "save_template", "Save template", class_="btn btn-sm btn-outline-success"
                ),

                ui.hr(style="margin-top: 8px; margin-bottom: 8px;"),
                ui.h4("Filters"),
                ui.input_selectize(
                    "summary", "Summary Type:",
                    choices=config.SUMMARY_TYPES, selected="Game Average",
                ),
                ui.input_slider("min_games", "Minimum Games:", 1, 10, 1, step=1),
                ui.input_checkbox_group(
                    "game_types", None,
                    choices=["Regular", "Finals"], selected=["Regular"], inline=True,
                ),
                ui.hr(style="margin-top: 0px; margin-bottom: 0px;"),
                ui.input_selectize("team", "Teams:", choices=[], selected=[], multiple=True),
                ui.input_selectize("player", "Players:", choices=[], selected=[], multiple=True),
                ui.input_selectize(
                    "position", "Positions:",
                    choices=config.POSITIONS, selected=[], multiple=True,
                ),
                ui.input_selectize(
                    "contract_end", "Contract Ends:",
                    choices=contract_end_choices, selected=[], multiple=True,
                ),
                ui.input_checkbox("position_separate", "Separate Positions", value=False),
                ui.input_checkbox("season_separate", "Separate Seasons", value=False),
                ui.input_checkbox("comp_separate", "Separate Comps", value=False),
                ui.hr(style="margin-top: 0px; margin-bottom: 0px;"),
                ui.input_selectize(
                    "stats", "Stats:",
                    choices=config.STAT_CHOICES, selected=config.DEFAULT_STATS, multiple=True,
                ),
                ui.hr(style="margin-top: 0px; margin-bottom: 0px;"),
                ui.h6("Highlight Rules"),
                ui.output_ui("highlight_rules_ui"),
                ui.input_action_button(
                    "add_rule", "Add rule", class_="btn btn-sm btn-outline-secondary mt-1"
                ),
                width=450,
            ),
            ui.tags.style(_table_css(resolve_id("table"))),
            toolbar(
                ui.tooltip(
                    ui.download_button(
                        "download_table",
                        None,
                        icon=DOWNLOAD_ICON,
                        class_="btn btn-sm btn-outline-success",
                        # Icon-only, so the button needs an explicit name for
                        # screen readers.
                        **{"aria-label": "Download Excel"},
                    ),
                    "Download Excel",
                ),
            ),
            ui.div("Tip: click a player row to open their profile.",
                   class_="pp-caption", style="margin:2px 2px 4px;"),
            ui.output_data_frame("table"),
        ),
    )


@module.server
def player_table_server(input, output, session, bigquery_data, contracts_df,
                        credentials, nav_request=None):

    templates = reactive.value(queries.fetch_player_table_templates(credentials))
    highlight_rules = reactive.value([])
    # Bumped on each row click so a repeat click on the same player still fires
    # a distinct nav request.
    _select_seq = reactive.value(0)

    # -------------------------
    # Filtering and summarising
    # -------------------------

    @reactive.effect
    def update_team_choices():
        df = bigquery_data()
        teams = (
            df[["teamId", "teamNickName"]]
            .drop_duplicates()
            .sort_values("teamNickName")
            .set_index("teamId")["teamNickName"]
            .astype(str)
            .to_dict()
        )
        ui.update_selectize("team", choices={str(k): v for k, v in teams.items()})

    @reactive.effect
    def update_player_choices():
        df = bigquery_data()
        selected_teams = input.team()

        if selected_teams:
            df = df[df["teamId"].astype(str).isin(selected_teams)]

        players = (
            df[["playerId", "playerName"]]
            .drop_duplicates()
            .sort_values("playerName")
            .set_index("playerId")["playerName"]
            .astype(str)
            .to_dict()
        )
        ui.update_selectize("player", choices={str(k): v for k, v in players.items()})

    @reactive.calc
    def filtered_data():
        teams = input.team()
        players = input.player()

        stats = list(config.STATS_DICT["Always"].keys()) + list(input.stats())

        return processing.filter_bq_player_data(
            bigquery_data(),
            input.game_types(),
            [int(t) for t in teams] if teams else None,
            [int(p) for p in players] if players else None,
            input.position(),
            stats,
        )

    @reactive.calc
    def summarised_data():
        summarised_df = processing.summarise_filtered_data(
            filtered_data(),
            input.summary(),
            input.min_games(),
            input.position_separate(),
            input.season_separate(),
            input.comp_separate(),
            ["mins"] + list(input.stats()),
            config.STATS_FLAT,
        )
        summarised_df = processing.add_contract_info(summarised_df, contracts_df)

        # Contract info only exists post-merge, so this filter applies last.
        return processing.filter_by_contract_end(
            summarised_df, input.contract_end(), config.UNSIGNED_LABEL
        )

    @reactive.calc
    def display_table_data():
        df = summarised_data().reset_index(drop=True)
        return df.drop(columns=["PID", "all_contract_end"], errors="ignore")

    @reactive.calc
    def scale_by_matches():
        return input.summary() in config.TOTALS_SUMMARY_TYPES

    # -------------------------
    # Highlight rules
    # -------------------------

    @reactive.calc
    def numeric_columns():
        return display_table_data().select_dtypes(include="number").columns.tolist()

    @reactive.effect
    @reactive.event(input.add_rule)
    def add_highlight_rule():
        numeric_cols = numeric_columns()
        if not numeric_cols:
            return

        highlight_rules.set(highlight_rules() + [{
            "id": str(uuid.uuid4())[:8],
            "col": numeric_cols[0],
            "op": ">",
            "val": 0.0,
            "color": next(iter(config.HIGHLIGHT_COLORS)),
        }])

    @render.ui
    def highlight_rules_ui():
        numeric_cols = numeric_columns()
        if not numeric_cols:
            return ui.p("No numeric columns available to highlight.")

        rows = []
        for rule in highlight_rules():
            rid = rule["id"]
            rows.append(
                ui.div(
                    {"style": "display:flex; gap:4px; align-items:center; margin-bottom:4px;"},
                    ui.input_select(
                        f"rule_col_{rid}", None,
                        choices=numeric_cols,
                        selected=rule["col"] if rule["col"] in numeric_cols else numeric_cols[0],
                        width="140px",
                    ),
                    ui.input_select(
                        f"rule_op_{rid}", None,
                        choices=RULE_OPERATORS, selected=rule["op"], width="55px",
                    ),
                    ui.input_numeric(f"rule_val_{rid}", None, value=rule["val"], width="70px"),
                    ui.input_select(
                        f"rule_color_{rid}", None,
                        choices=config.HIGHLIGHT_COLORS, selected=rule["color"], width="80px",
                    ),
                    ui.input_action_button(
                        f"rule_delete_{rid}", "✕",
                        class_="btn btn-sm btn-outline-danger", style="padding:2px 7px;",
                    ),
                )
            )

        return ui.div(*rows)

    @reactive.effect
    def sync_and_delete_highlight_rules():
        """Pull rule-row input values back into the rules list, dropping deleted rows."""
        new_rules = []
        changed = False

        for rule in highlight_rules():
            rid = rule["id"]

            # The inputs for a rule only exist once its row has rendered, so a
            # missing input means "leave this rule as it is".
            try:
                delete_count = input[f"rule_delete_{rid}"]()
            except Exception:
                delete_count = 0

            if delete_count > 0:
                changed = True
                continue

            try:
                updated_rule = {
                    "id": rid,
                    "col": input[f"rule_col_{rid}"](),
                    "op": input[f"rule_op_{rid}"](),
                    "val": input[f"rule_val_{rid}"](),
                    "color": input[f"rule_color_{rid}"](),
                }
            except Exception:
                new_rules.append(rule)
                continue

            if any(v is None for v in updated_rule.values()):
                new_rules.append(rule)
                continue

            updated_rule["val"] = float(updated_rule["val"])

            if updated_rule != rule:
                changed = True

            new_rules.append(updated_rule)

        if changed:
            highlight_rules.set(new_rules)

    # -------------------------
    # Templates
    # -------------------------

    @reactive.effect
    def update_template_choices():
        ui.update_selectize(
            "template",
            choices=list(templates().keys()),
            selected=input.template(),
        )

    @reactive.effect
    @reactive.event(input.save_template)
    def save_current_template():
        template_name = input.new_template_name()
        if not template_name:
            return

        saved = dict(templates())
        saved[template_name] = {
            "summary": input.summary(),
            "min_games": input.min_games(),
            "game_types": list(input.game_types()),
            "team": list(input.team()),
            "player": list(input.player()),
            "position": list(input.position()),
            "contract_end": list(input.contract_end()),
            "position_separate": input.position_separate(),
            "season_separate": input.season_separate(),
            "comp_separate": input.comp_separate(),
            "stats": list(input.stats()),
            "highlight_rules": [
                {k: rule[k] for k in ("col", "op", "val", "color")}
                for rule in highlight_rules()
            ],
        }

        queries.save_player_table_templates(credentials, saved)
        templates.set(saved)

        ui.update_selectize("template", choices=list(saved.keys()), selected=template_name)
        ui.update_text("new_template_name", value="")

    def apply_template_to_inputs(template):
        ui.update_selectize("summary", selected=template.get("summary", "Game Average"))
        ui.update_slider("min_games", value=template.get("min_games", 1))
        ui.update_checkbox_group("game_types", selected=template.get("game_types", ["Regular"]))
        ui.update_selectize("team", selected=template.get("team", []))
        ui.update_selectize("player", selected=template.get("player", []))
        ui.update_selectize("position", selected=template.get("position", []))
        # Templates saved before this filter existed have no key, so default
        # to "no contract filter" rather than dropping every player.
        ui.update_selectize("contract_end", selected=template.get("contract_end", []))
        ui.update_checkbox("position_separate", value=template.get("position_separate", False))
        ui.update_checkbox("season_separate", value=template.get("season_separate", False))
        ui.update_checkbox("comp_separate", value=template.get("comp_separate", False))
        ui.update_selectize("stats", selected=template.get("stats", config.DEFAULT_STATS))

        highlight_rules.set([
            {"id": str(uuid.uuid4())[:8], **{k: rule[k] for k in ("col", "op", "val", "color")}}
            for rule in template.get("highlight_rules", [])
        ])

    @reactive.effect
    @reactive.event(input.apply_template)
    def apply_selected_template():
        template_name = input.template()
        saved = templates()

        if template_name and template_name in saved:
            apply_template_to_inputs(saved[template_name])

    default_applied = reactive.value(False)

    @reactive.effect
    def apply_default_on_launch():
        if default_applied():
            return

        saved = templates()
        if "Default" in saved:
            apply_template_to_inputs(saved["Default"])
            ui.update_selectize("template", selected="Default")

        default_applied.set(True)

    @reactive.effect
    @reactive.event(input.delete_template)
    def delete_selected_template():
        template_name = input.template()
        if not template_name:
            return

        saved = dict(templates())
        saved.pop(template_name, None)

        queries.save_player_table_templates(credentials, saved)
        templates.set(saved)

        ui.update_selectize("template", choices=list(saved.keys()), selected=None)

    # -------------------------
    # Table and export
    # -------------------------

    def contract_styles(df, display_df):
        """Colour player names by contract expiry year."""
        if "all_contract_end" not in df.columns or "Name" not in display_df.columns:
            return []

        styles = []
        for year, color in config.CONTRACT_END_COLORS.items():
            rows = df.index[df["all_contract_end"] == year].tolist()
            if rows:
                styles.append(render.StyleInfo(
                    rows=rows, cols=["Name"],
                    style={"color": color, "font-weight": "bold"},
                ))

        unsigned_rows = df.index[df["all_contract_end"].isna()].tolist()
        if unsigned_rows:
            styles.append(render.StyleInfo(
                rows=unsigned_rows, cols=["Name"],
                style={"color": config.UNSIGNED_COLOR, "font-weight": "bold"},
            ))

        return styles

    @render.data_frame
    def table():
        df = summarised_data().reset_index(drop=True)
        display_df = display_table_data()

        styles = contract_styles(df, display_df)

        for rule in highlight_rules():
            if rule["col"] not in display_df.columns:
                continue

            mask = highlight_mask(display_df, rule, scale_by_matches())
            rows = display_df.index[mask].tolist()

            if rows:
                styles.append(render.StyleInfo(
                    rows=rows,
                    cols=[rule["col"]],
                    style={
                        "background-color": processing.hex_to_rgba(rule["color"], 0.45),
                        "font-weight": "bold",
                    },
                ))

        return render.DataTable(display_df, styles=styles, width="100%", height="99%",
                                selection_mode="row")

    @reactive.effect
    def open_profile_on_row_click():
        """Selecting a row opens that player's profile (see app.py nav_request)."""
        if nav_request is None:
            return
        selection = table.cell_selection()
        rows = (selection or {}).get("rows") or ()
        if not rows:
            return
        df = summarised_data().reset_index(drop=True)
        idx = rows[0]
        if "PID" not in df.columns or idx >= len(df):
            return
        pid = int(df.iloc[idx]["PID"])
        name = str(df.iloc[idx]["Name"]) if "Name" in df.columns else str(pid)
        with reactive.isolate():
            seq = _select_seq() + 1
        _select_seq.set(seq)
        nav_request.set((pid, name, seq))

    @render.download(filename="player_table.xlsx")
    def download_table():
        yield build_workbook(
            summarised_data().reset_index(drop=True),
            display_table_data(),
            highlight_rules(),
            scale_by_matches(),
        )
